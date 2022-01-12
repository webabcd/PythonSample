import openpyxl

path = r'D:\gitroot\PythonSample\basic\test.xlsx'
workbook = openpyxl.load_workbook(path)
sheets = workbook.sheetnames

sheet_new = workbook.create_sheet('new_sheet')
i_new = 0

for sheet_index in range(len(sheets)):
    sheet = workbook[sheets[sheet_index]]
    for i in range(sheet.max_row):
        for j in range(sheet.max_column):
            sheet_new.cell(i_new + 1, j + 1).value = sheet.cell(i + 1, j + 1).value            
        i_new += 1

workbook.save(path)